"""Excel (.xlsx) writer, via openpyxl.

Numbers and dates are written as native cell types rather than strings, so Excel
can sort, filter and format them — writing "4.6" as text produces a workbook
where sorting by lead score is alphabetical, which is the usual complaint about
machine-generated spreadsheets. Only genuinely textual cells go through formula
neutralization; a native numeric cell cannot be a formula.

Reports become one sheet per section, which is what makes XLSX the best format
for the dashboard and analytics exports.

`write_only` mode is deliberately *not* used: it would lower peak memory for
very large sheets, but it rules out freeze panes, auto-filter and column sizing
after the fact. Large exports are bounded by `EXPORT_MAX_ROWS` and run in a
Celery worker, so the formatting is worth more here than the memory.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from services.exporters.dataset import Column, Dataset, ReportSection, format_cell, neutralize_formula

_HEADER_FILL = PatternFill("solid", start_color="FF1F2937", end_color="FF1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14)
_META_KEY_FONT = Font(bold=True, size=9, color="FF6B7280")
_META_VALUE_FONT = Font(size=9, color="FF6B7280")

# Excel hard limits. Sheet titles are capped at 31 chars and may not contain
# these characters; exceeding either makes the workbook unopenable.
_MAX_SHEET_TITLE = 31
_INVALID_SHEET_CHARS = str.maketrans({c: "-" for c in r"[]:*?/\\"})

# Excel refuses to open a file containing a cell longer than this.
_MAX_CELL_CHARS = 32_767


def render(dataset: Dataset) -> bytes:
    workbook = Workbook()
    # Workbook() starts with one sheet; reuse it for the first table.
    first_sheet = workbook.active

    if dataset.columns:
        first_sheet.title = _sheet_title(dataset.title)
        _write_sheet(first_sheet, dataset, dataset.columns, dataset.rows, include_header_block=True)
        extra_sections = dataset.sections
    elif dataset.sections:
        head, *rest = dataset.sections
        first_sheet.title = _sheet_title(head.title)
        _write_sheet(first_sheet, dataset, head.columns, head.rows, include_header_block=True, section=head)
        extra_sections = rest
    else:
        # No columns and no sections: still produce a valid, self-describing
        # workbook rather than a corrupt zero-sheet file.
        first_sheet.title = _sheet_title(dataset.title)
        _write_sheet(first_sheet, dataset, [], [], include_header_block=True)
        extra_sections = []

    for section in extra_sections:
        sheet = workbook.create_sheet(_sheet_title(section.title, workbook))
        _write_sheet(sheet, dataset, section.columns, section.rows, include_header_block=False, section=section)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _write_sheet(
    sheet: Worksheet,
    dataset: Dataset,
    columns: list[Column],
    rows: list[dict],
    *,
    include_header_block: bool,
    section: ReportSection | None = None,
) -> None:
    cursor = 1

    if include_header_block:
        sheet.cell(row=cursor, column=1, value=dataset.title).font = _TITLE_FONT
        cursor += 1
        if dataset.subtitle:
            sheet.cell(row=cursor, column=1, value=dataset.subtitle).font = _META_VALUE_FONT
            cursor += 1
        for key, value in dataset.metadata.items():
            sheet.cell(row=cursor, column=1, value=key).font = _META_KEY_FONT
            sheet.cell(row=cursor, column=2, value=_coerce(value)).font = _META_VALUE_FONT
            cursor += 1
        cursor += 1  # spacer row

    if section is not None:
        sheet.cell(row=cursor, column=1, value=section.title).font = _TITLE_FONT
        cursor += 1
        if section.note:
            sheet.cell(row=cursor, column=1, value=section.note).font = _META_VALUE_FONT
            cursor += 1

    if not columns:
        return

    header_row = cursor
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column.label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = column.width
    sheet.row_dimensions[header_row].height = 20

    for offset, row in enumerate(rows, start=1):
        for index, column in enumerate(columns, start=1):
            sheet.cell(row=header_row + offset, column=index, value=_coerce(row.get(column.key)))

    last_row = header_row + len(rows)
    # Freeze below the header so it stays visible while scrolling, and add an
    # auto-filter so the recipient can slice the export without a pivot table.
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    if rows:
        sheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(columns))}{last_row}"
        )


def _coerce(value: object):
    """Keeps numbers/dates native; neutralizes and truncates text."""
    if value is None:
        return None
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, Decimal):
        # openpyxl cannot serialize Decimal; float keeps it numeric in Excel.
        return float(value)
    if isinstance(value, datetime):
        # openpyxl cannot write a tz-aware datetime ("cannot be represented in
        # Excel"). Excel has no timezone concept, so normalize to naive UTC.
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, date):
        return value
    return neutralize_formula(format_cell(value))[:_MAX_CELL_CHARS]


def _sheet_title(raw: str, workbook: Workbook | None = None) -> str:
    title = (raw or "Sheet").translate(_INVALID_SHEET_CHARS).strip() or "Sheet"
    title = title[:_MAX_SHEET_TITLE]
    if workbook is None:
        return title

    # Sheet names must be unique; two analytics sections could collide after
    # truncation, which openpyxl would silently rename or reject.
    existing = set(workbook.sheetnames)
    if title not in existing:
        return title
    for suffix in range(2, 100):
        candidate = f"{title[: _MAX_SHEET_TITLE - len(str(suffix)) - 1]}-{suffix}"
        if candidate not in existing:
            return candidate
    return title[:_MAX_SHEET_TITLE]
