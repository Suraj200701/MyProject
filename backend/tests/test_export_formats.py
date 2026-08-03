"""Unit tests for the export writers and the `Dataset` contract.

No database, no HTTP, no event loop — these render bytes and parse them back, so
they run in about a second and pin down writer behaviour that would otherwise
only be exercised indirectly through the API (see tests/test_exports.py).

The security-relevant assertions here are the formula-injection ones: lead data
is attacker-influenced (company names arrive from third-party providers and from
user-supplied CSV imports), so a cell that a recipient's spreadsheet would
evaluate is a real vulnerability, not a formatting nit.
"""

import csv
import io
import json
from datetime import UTC, date, datetime
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from models.enums import ExportFormat
from schemas.export import format_size
from services import exporters
from services.exporters.dataset import (
    Column,
    Dataset,
    ReportSection,
    format_cell,
    neutralize_formula,
    text_cell,
)

COLUMNS = [
    Column("company", "Company", 30),
    Column("score", "Lead Score", 10),
    Column("rating", "Rating", 8),
    Column("created", "Created", 18),
    Column("tags", "Tags", 20),
    Column("verified", "Verified", 8),
]

ROWS = [
    {
        "company": "Apex Switchgear Pvt Ltd", "score": 87, "rating": Decimal("4.6"),
        "created": datetime(2026, 7, 30, 12, 0, tzinfo=UTC),
        "tags": ["Electrical", "MSME"], "verified": True,
    },
    {"company": "Nova Panels", "score": 12, "rating": None, "created": None, "tags": [], "verified": False},
]


def table() -> Dataset:
    return Dataset(
        title="Leads Export",
        subtitle="2 leads",
        columns=COLUMNS,
        rows=list(ROWS),
        metadata={"Organization": "Test Org", "Scope": "All leads"},
    )


def report() -> Dataset:
    return Dataset(
        title="Analytics Report",
        metadata={"Organization": "Test Org"},
        sections=[
            ReportSection(
                "Top Industries",
                [Column("name", "Industry", 25), Column("n", "Leads", 10)],
                [{"name": "Electrical", "n": 120}, {"name": "EPC", "n": 64}],
                note="Last 90 days",
            ),
            ReportSection(
                "Top Cities",
                [Column("city", "City", 25), Column("n", "Leads", 10)],
                [{"city": "Pune", "n": 88}],
            ),
        ],
    )


# --- Cell coercion --------------------------------------------------------


@pytest.mark.parametrize(
    "value,expected",
    [
        (None, ""),
        (True, "Yes"),
        (False, "No"),
        (42, "42"),
        (Decimal("4.0"), "4"),          # a rating of 4 should not read as "4.0"
        (Decimal("4.6"), "4.6"),
        (4.0, "4"),
        (4.5, "4.5"),
        (date(2026, 7, 30), "2026-07-30"),
        (datetime(2026, 7, 30, 12, 0), "2026-07-30 12:00:00"),
        (["a", "b"], "a, b"),
        ([None, "a"], "a"),
        ({"k": 1}, "k=1"),
        ("plain", "plain"),
    ],
)
def test_format_cell(value, expected):
    assert format_cell(value) == expected


@pytest.mark.parametrize("payload", ["=1+1", "+1", "-1", "@SUM(A1)", "\tx", "\rx"])
def test_formula_triggers_are_neutralized(payload):
    assert neutralize_formula(payload) == f"'{payload}"


@pytest.mark.parametrize("safe", ["Apex Ltd", "4.6", "sales@apex.com", "", "Smith & Sons"])
def test_safe_values_are_left_alone(safe):
    assert neutralize_formula(safe) == safe


def test_text_cell_formats_then_neutralizes():
    assert text_cell(["=evil", "b"]) == "'=evil, b"


def test_row_count_spans_sections():
    """The row cap must see report sections, not just the primary table."""
    assert report().row_count == 3
    assert table().row_count == 2


# --- CSV ------------------------------------------------------------------


def test_csv_has_a_bom_for_excel():
    blob = exporters.render(table(), ExportFormat.CSV)
    assert blob.startswith(b"\xef\xbb\xbf")


def test_csv_uses_crlf_line_endings():
    """RFC 4180, and what Excel expects."""
    blob = exporters.render(table(), ExportFormat.CSV)
    assert b"\r\n" in blob


def test_csv_round_trips_through_the_csv_module():
    blob = exporters.render(table(), ExportFormat.CSV)
    rows = list(csv.reader(io.StringIO(blob.decode("utf-8-sig"))))
    header = next(r for r in rows if r and r[0] == "Company")
    assert header == [c.label for c in COLUMNS]
    data = rows[rows.index(header) + 1]
    assert data == ["Apex Switchgear Pvt Ltd", "87", "4.6", "2026-07-30 12:00:00", "Electrical, MSME", "Yes"]


def test_csv_carries_the_metadata_block():
    blob = exporters.render(table(), ExportFormat.CSV).decode("utf-8-sig")
    assert "Leads Export" in blob
    assert "Test Org" in blob
    assert "All leads" in blob


def test_csv_neutralizes_injection():
    dataset = table()
    dataset.rows.append({**ROWS[1], "company": "=cmd|'/c calc'!A1"})
    blob = exporters.render(dataset, ExportFormat.CSV).decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(blob)))
    assert any(r and r[0] == "'=cmd|'/c calc'!A1" for r in rows)


def test_csv_flattens_report_sections():
    blob = exporters.render(report(), ExportFormat.CSV).decode("utf-8-sig")
    assert "Top Industries" in blob
    assert "Top Cities" in blob
    assert "Last 90 days" in blob


def test_csv_quotes_embedded_delimiters():
    dataset = Dataset(
        title="t", columns=[Column("a", "A")],
        rows=[{"a": 'has, comma and "quote"'}, {"a": "has\nnewline"}],
    )
    blob = exporters.render(dataset, ExportFormat.CSV).decode("utf-8-sig")
    parsed = list(csv.reader(io.StringIO(blob)))
    values = [r[0] for r in parsed if r]
    assert 'has, comma and "quote"' in values
    assert "has\nnewline" in values


# --- XLSX -----------------------------------------------------------------


def test_xlsx_opens_as_a_workbook():
    workbook = load_workbook(io.BytesIO(exporters.render(table(), ExportFormat.EXCEL)))
    assert workbook.sheetnames


def test_xlsx_keeps_numbers_and_dates_native():
    """Text-typed numbers make Excel sort lead scores alphabetically."""
    sheet = load_workbook(io.BytesIO(exporters.render(table(), ExportFormat.EXCEL))).active
    values = [[c.value for c in row] for row in sheet.iter_rows()]
    header = next(r for r in values if r and r[0] == "Company")
    data = values[values.index(header) + 1]

    assert isinstance(data[1], int)                    # score
    assert isinstance(data[2], float)                  # rating (Decimal -> float)
    assert isinstance(data[3], datetime)               # created
    assert data[3].tzinfo is None, "Excel has no timezone concept"


def test_xlsx_has_a_frozen_header_and_autofilter():
    sheet = load_workbook(io.BytesIO(exporters.render(table(), ExportFormat.EXCEL))).active
    assert sheet.freeze_panes is not None
    assert sheet.auto_filter.ref is not None


def test_xlsx_neutralizes_injection():
    dataset = table()
    dataset.rows.append({**ROWS[1], "company": "=HYPERLINK('http://evil','x')"})
    sheet = load_workbook(io.BytesIO(exporters.render(dataset, ExportFormat.EXCEL))).active
    strings = [c.value for row in sheet.iter_rows() for c in row if isinstance(c.value, str)]
    assert any(v.startswith("'=") for v in strings)


def test_xlsx_gives_each_report_section_its_own_sheet():
    workbook = load_workbook(io.BytesIO(exporters.render(report(), ExportFormat.EXCEL)))
    assert workbook.sheetnames == ["Top Industries", "Top Cities"]


def test_xlsx_sanitizes_illegal_sheet_names():
    """Excel refuses to open a workbook with []:*?/\\ in a sheet name."""
    dataset = Dataset(
        title="t",
        sections=[ReportSection("Bad/Name:[x]*?", [Column("a", "A")], [{"a": 1}])],
    )
    workbook = load_workbook(io.BytesIO(exporters.render(dataset, ExportFormat.EXCEL)))
    assert not any(ch in workbook.sheetnames[0] for ch in r"[]:*?/\\")


def test_xlsx_truncates_overlong_sheet_names():
    """Excel caps sheet titles at 31 characters."""
    dataset = Dataset(title="t", sections=[ReportSection("x" * 60, [Column("a", "A")], [{"a": 1}])])
    workbook = load_workbook(io.BytesIO(exporters.render(dataset, ExportFormat.EXCEL)))
    assert len(workbook.sheetnames[0]) <= 31


def test_xlsx_deduplicates_colliding_sheet_names():
    """Two sections can collide after truncation; sheet names must stay unique."""
    long_name = "y" * 40
    dataset = Dataset(
        title="t",
        sections=[
            ReportSection(long_name, [Column("a", "A")], [{"a": 1}]),
            ReportSection(long_name, [Column("a", "A")], [{"a": 2}]),
        ],
    )
    workbook = load_workbook(io.BytesIO(exporters.render(dataset, ExportFormat.EXCEL)))
    assert len(set(workbook.sheetnames)) == len(workbook.sheetnames) == 2


# --- PDF ------------------------------------------------------------------


def test_pdf_is_structurally_valid():
    blob = exporters.render(table(), ExportFormat.PDF)
    assert blob.startswith(b"%PDF-")
    assert b"%%EOF" in blob[-1024:]


def test_pdf_escapes_reportlab_markup():
    """Unescaped <, > or & in real company names fails the whole build."""
    dataset = table()
    dataset.rows.append({**ROWS[1], "company": "Smith & Sons <Holdings> Ltd"})
    assert exporters.render(dataset, ExportFormat.PDF).startswith(b"%PDF-")


def test_pdf_renders_report_sections():
    assert exporters.render(report(), ExportFormat.PDF).startswith(b"%PDF-")


def test_pdf_handles_an_empty_dataset():
    assert exporters.render(Dataset(title="Nothing"), ExportFormat.PDF).startswith(b"%PDF-")


def test_pdf_caps_very_large_tables():
    """PDF is a presentation format; past the cap it says so rather than truncating silently."""
    from services.exporters import pdf_exporter

    dataset = Dataset(
        title="Big",
        columns=[Column("a", "A", 20)],
        rows=[{"a": f"row-{i}"} for i in range(pdf_exporter._MAX_TABLE_ROWS + 50)],
    )
    blob = exporters.render(dataset, ExportFormat.PDF)
    assert blob.startswith(b"%PDF-")


# --- JSON -----------------------------------------------------------------


def test_json_preserves_types():
    payload = json.loads(exporters.render(table(), ExportFormat.JSON))
    row = payload["rows"][0]
    assert row["score"] == 87 and isinstance(row["score"], int)
    assert row["rating"] == 4.6
    assert row["verified"] is True
    assert row["tags"] == ["Electrical", "MSME"]
    assert row["created"].startswith("2026-07-30T12:00:00")


def test_json_does_not_neutralize_formulas():
    """Nothing evaluates a JSON string; escaping would corrupt the value."""
    dataset = table()
    dataset.rows.append({**ROWS[1], "company": "=1+1"})
    payload = json.loads(exporters.render(dataset, ExportFormat.JSON))
    assert any(r["company"] == "=1+1" for r in payload["rows"])


def test_json_projects_only_the_selected_columns():
    dataset = Dataset(title="t", columns=[Column("company", "Company")], rows=list(ROWS))
    payload = json.loads(exporters.render(dataset, ExportFormat.JSON))
    assert set(payload["rows"][0]) == {"company"}


def test_json_includes_report_sections_and_row_count():
    payload = json.loads(exporters.render(report(), ExportFormat.JSON))
    assert [s["title"] for s in payload["sections"]] == ["Top Industries", "Top Cities"]
    assert payload["row_count"] == 3


# --- Dispatch and file naming --------------------------------------------


@pytest.mark.parametrize("fmt", list(ExportFormat))
def test_every_enum_member_has_a_writer(fmt):
    """A new format must not be addable to the enum without a writer."""
    assert exporters.render(table(), fmt)


@pytest.mark.parametrize(
    "fmt,extension",
    [(ExportFormat.CSV, "csv"), (ExportFormat.EXCEL, "xlsx"), (ExportFormat.PDF, "pdf"), (ExportFormat.JSON, "json")],
)
def test_extensions(fmt, extension):
    assert exporters.extension_for(fmt) == extension


@pytest.mark.parametrize(
    "stem,fmt,expected",
    [
        ("leads", ExportFormat.CSV, "leads.csv"),
        ("leads.csv", ExportFormat.CSV, "leads.csv"),          # already correct
        ("leads.xls", ExportFormat.EXCEL, "leads.xlsx"),       # wrong guess corrected
        ("report.csv", ExportFormat.EXCEL, "report.xlsx"),     # bytes win over the name
        ("", ExportFormat.CSV, "export.csv"),
        (None, ExportFormat.PDF, "export.pdf"),
        ("my.company.leads", ExportFormat.CSV, "my.company.leads.csv"),
    ],
)
def test_build_file_name(stem, fmt, expected):
    assert exporters.build_file_name(stem, fmt) == expected


def test_media_types_are_distinct_per_format():
    types = {exporters.media_type_for(f) for f in ExportFormat}
    assert len(types) == len(list(ExportFormat))


# --- Size label -----------------------------------------------------------


@pytest.mark.parametrize(
    "size,expected",
    [
        (0, "0 B"), (512, "512 B"), (1024, "1 KB"), (410 * 1024, "410 KB"),
        (1024 * 1024, "1.0 MB"), (int(1.25 * 1024 * 1024), "1.2 MB"), (2 * 1024**3, "2.0 GB"),
    ],
)
def test_size_label_formatting(size, expected):
    assert format_size(size) == expected
