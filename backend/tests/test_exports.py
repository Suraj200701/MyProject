"""Export Center integration tests.

Driven through the HTTP API against the real database, real storage backend and
real file writers — every generated file is parsed back with the library that
would open it (`csv`, `openpyxl`, `json`, and the PDF magic bytes), because an
export that returns 201 and produces a corrupt file is the failure mode that
matters here.

Covered: all four formats, all four resources, all three scopes, column
selection, history pagination and filters, both download paths, and the security
requirements (tenant isolation, RBAC, rate limiting, row/size caps, expiry,
cleanup).
"""

import csv
import io
import json
import time
import uuid
from datetime import UTC, datetime, timedelta

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select

from models.enums import ExportFormat, ExportResource, ExportStatus, RoleName
from models.organization import Organization, OrganizationMember
from models.search import Export
from models.user import Role

pytestmark = pytest.mark.asyncio(loop_scope="session")


@pytest.fixture(autouse=True)
def isolated_upload_dir(tmp_path, monkeypatch):
    """Points storage at a per-test temp dir.

    Keeps generated files out of the repo's uploads/ directory and guarantees the
    cleanup assertions are looking at files this test created.
    """
    from config.settings import settings

    monkeypatch.setattr(settings, "UPLOAD_DIR", str(tmp_path / "uploads"), raising=False)


async def _create_leads(client: AsyncClient, headers: dict, specs: list[dict]) -> list[str]:
    ids = []
    for spec in specs:
        resp = await client.post("/api/v1/leads", headers=headers, json=spec)
        assert resp.status_code == 201, resp.text
        ids.append(resp.json()["id"])
    return ids


LEAD_SPECS = [
    {
        "company": "Apex Switchgear Pvt Ltd", "industry": "Electrical", "city": "Pune",
        "country": "India", "contact_name": "Rohan Desai", "email": "sales@apexswitchgear.com",
        "phone": "+919876543210", "website": "https://apexswitchgear.com",
        "gst_number": "27AAPFU0939F1ZV", "lead_score": 88, "status": "new",
    },
    {
        "company": "Nova Control Panels", "industry": "Electrical", "city": "Thane",
        "country": "India", "contact_name": "Priya Nair", "email": "info@novapanels.co.in",
        "phone": "+912245678901", "lead_score": 64, "status": "contacted",
    },
    {
        "company": "Zenith Textiles Ltd", "industry": "Textiles", "city": "Surat",
        "country": "India", "email": "hello@zenithtex.in", "lead_score": 35, "status": "new",
    },
]


@pytest.fixture
async def seeded_leads(client: AsyncClient, signed_up_user):
    _, headers = signed_up_user
    ids = await _create_leads(client, headers, LEAD_SPECS)
    return headers, ids


# --- Format coverage ------------------------------------------------------


@pytest.mark.parametrize(
    "fmt,extension",
    [("csv", "csv"), ("excel", "xlsx"), ("pdf", "pdf"), ("json", "json")],
)
async def test_every_format_produces_a_ready_downloadable_export(client, seeded_leads, fmt, extension):
    headers, _ = seeded_leads

    resp = await client.post("/api/v1/exports", headers=headers, json={"format": fmt, "file_name": "my_leads"})
    assert resp.status_code == 201, resp.text
    body = resp.json()

    assert body["status"] == "ready"
    assert body["format"] == fmt
    assert body["resource"] == "leads"
    assert body["file_name"] == f"my_leads.{extension}"
    assert body["row_count"] == 3
    assert body["size_bytes"] > 0
    assert body["size_label"]
    assert body["download_url"] == f"/api/v1/exports/{body['id']}/download"
    assert body["expires_at"] is not None

    download = await client.get(body["download_url"], headers=headers)
    assert download.status_code == 200
    assert len(download.content) == body["size_bytes"]
    assert f'filename="my_leads.{extension}"' in download.headers["content-disposition"]
    assert download.headers["cache-control"] == "private, no-store"
    assert download.headers["x-content-type-options"] == "nosniff"


async def test_csv_export_parses_and_contains_the_leads(client, seeded_leads):
    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    blob = (await client.get(created.json()["download_url"], headers=headers)).content

    assert blob.startswith(b"\xef\xbb\xbf"), "Excel needs the UTF-8 BOM to decode this correctly"
    rows = list(csv.reader(io.StringIO(blob.decode("utf-8-sig"))))
    header_index = next(i for i, r in enumerate(rows) if r and r[0] == "Company")
    companies = {r[0] for r in rows[header_index + 1 :] if r}
    assert companies == {"Apex Switchgear Pvt Ltd", "Nova Control Panels", "Zenith Textiles Ltd"}


async def test_excel_export_opens_as_a_workbook(client, seeded_leads):
    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "excel"})
    blob = (await client.get(created.json()["download_url"], headers=headers)).content

    workbook = load_workbook(io.BytesIO(blob))
    sheet = workbook.active
    values = [[c.value for c in row] for row in sheet.iter_rows()]
    header_index = next(i for i, r in enumerate(values) if r and r[0] == "Company")
    assert sheet.freeze_panes is not None
    # Lead score must be a real number so Excel sorts it numerically.
    score_column = values[header_index].index("Lead Score")
    assert isinstance(values[header_index + 1][score_column], int)


async def test_pdf_export_is_a_valid_pdf(client, seeded_leads):
    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "pdf"})
    blob = (await client.get(created.json()["download_url"], headers=headers)).content
    assert blob.startswith(b"%PDF-")
    assert b"%%EOF" in blob[-1024:]


async def test_json_export_keeps_native_types(client, seeded_leads):
    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "json"})
    blob = (await client.get(created.json()["download_url"], headers=headers)).content

    payload = json.loads(blob)
    assert payload["row_count"] == 3
    scores = [r["lead_score"] for r in payload["rows"]]
    assert all(isinstance(s, int) for s in scores)


async def test_media_type_matches_the_format(client, seeded_leads):
    headers, _ = seeded_leads
    expected = {
        "csv": "text/csv",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "pdf": "application/pdf",
        "json": "application/json",
    }
    for fmt, media_type in expected.items():
        created = await client.post("/api/v1/exports", headers=headers, json={"format": fmt})
        download = await client.get(created.json()["download_url"], headers=headers)
        assert media_type in download.headers["content-type"], fmt


# --- Scopes ---------------------------------------------------------------


async def test_selected_scope_exports_only_the_chosen_leads(client, seeded_leads):
    headers, ids = seeded_leads

    resp = await client.post(
        "/api/v1/exports",
        headers=headers,
        json={"format": "csv", "scope": "selected", "lead_ids": [ids[0], ids[2]]},
    )
    assert resp.status_code == 201
    assert resp.json()["row_count"] == 2

    blob = (await client.get(resp.json()["download_url"], headers=headers)).content
    rows = list(csv.reader(io.StringIO(blob.decode("utf-8-sig"))))
    header_index = next(i for i, r in enumerate(rows) if r and r[0] == "Company")
    companies = {r[0] for r in rows[header_index + 1 :] if r}
    assert companies == {"Apex Switchgear Pvt Ltd", "Zenith Textiles Ltd"}


async def test_selected_scope_requires_lead_ids(client, seeded_leads):
    headers, _ = seeded_leads
    resp = await client.post("/api/v1/exports", headers=headers, json={"format": "csv", "scope": "selected"})
    assert resp.status_code == 400
    assert "lead_ids" in resp.json()["message"]


async def test_selected_scope_cannot_reach_another_orgs_leads(client, seeded_leads, db_session):
    """An id from another organization must contribute nothing, not leak a row."""
    headers, ids = seeded_leads
    other = await client.post(
        "/api/v1/auth/signup",
        json={
            "email": f"other_{uuid.uuid4().hex[:8]}@example.com", "password": "TestPass123",
            "full_name": "Other", "company_name": "Other Co",
        },
    )
    other_headers = {"Authorization": f"Bearer {other.json()['access_token']}"}
    foreign_ids = await _create_leads(other_headers and client or client, other_headers, [LEAD_SPECS[0]])

    resp = await client.post(
        "/api/v1/exports",
        headers=headers,
        json={"format": "csv", "scope": "selected", "lead_ids": [ids[0], foreign_ids[0]]},
    )
    assert resp.status_code == 201
    assert resp.json()["row_count"] == 1


async def test_filtered_scope_matches_what_the_leads_list_returns(client, seeded_leads):
    """The export and the table it came from must agree on the same filters."""
    headers, _ = seeded_leads
    filters = {"industry": "Electrical", "min_score": 60}

    listed = await client.get("/api/v1/leads", headers=headers, params=filters)
    expected = {item["company"] for item in listed.json()["items"]}
    assert expected == {"Apex Switchgear Pvt Ltd", "Nova Control Panels"}

    resp = await client.post(
        "/api/v1/exports", headers=headers, json={"format": "csv", "scope": "filtered", "filters": filters}
    )
    assert resp.json()["row_count"] == len(expected)

    blob = (await client.get(resp.json()["download_url"], headers=headers)).content
    rows = list(csv.reader(io.StringIO(blob.decode("utf-8-sig"))))
    header_index = next(i for i, r in enumerate(rows) if r and r[0] == "Company")
    assert {r[0] for r in rows[header_index + 1 :] if r} == expected


async def test_filter_summary_is_recorded_in_the_file(client, seeded_leads):
    """A file that left the product should say what selection produced it."""
    headers, _ = seeded_leads
    resp = await client.post(
        "/api/v1/exports",
        headers=headers,
        json={"format": "csv", "scope": "filtered", "filters": {"industry": "Textiles"}},
    )
    blob = (await client.get(resp.json()["download_url"], headers=headers)).content
    text = blob.decode("utf-8-sig")
    assert "Textiles" in text
    assert "Filters" in text


async def test_all_scope_exports_every_lead(client, seeded_leads):
    headers, _ = seeded_leads
    resp = await client.post("/api/v1/exports", headers=headers, json={"format": "csv", "scope": "all"})
    assert resp.json()["row_count"] == 3


# --- Column selection -----------------------------------------------------


async def test_columns_can_be_chosen_by_api_key(client, seeded_leads):
    headers, _ = seeded_leads
    resp = await client.post(
        "/api/v1/exports", headers=headers, json={"format": "csv", "columns": ["company", "email", "lead_score"]}
    )
    blob = (await client.get(resp.json()["download_url"], headers=headers)).content
    rows = list(csv.reader(io.StringIO(blob.decode("utf-8-sig"))))
    header = next(r for r in rows if r and r[0] == "Company")
    assert header == ["Company", "Email", "Lead Score"]


async def test_columns_accept_the_frontend_wizard_labels(client, seeded_leads):
    """The export wizard holds display labels in state; they must work as-is."""
    headers, _ = seeded_leads
    wizard_fields = ["Company", "Industry", "City", "Contact", "Email", "Phone", "Lead Score", "Status"]

    resp = await client.post("/api/v1/exports", headers=headers, json={"format": "csv", "columns": wizard_fields})
    assert resp.status_code == 201
    assert resp.json()["ignored_columns"] == []

    blob = (await client.get(resp.json()["download_url"], headers=headers)).content
    rows = list(csv.reader(io.StringIO(blob.decode("utf-8-sig"))))
    assert next(r for r in rows if r and r[0] == "Company") == wizard_fields


async def test_unknown_columns_are_reported_not_fatal(client, seeded_leads):
    headers, _ = seeded_leads
    resp = await client.post(
        "/api/v1/exports", headers=headers, json={"format": "csv", "columns": ["Company", "made_up_field"]}
    )
    assert resp.status_code == 201
    assert resp.json()["ignored_columns"] == ["made_up_field"]


async def test_json_export_omits_unselected_columns(client, seeded_leads):
    """Column choice must actually withhold data, not just hide it visually."""
    headers, _ = seeded_leads
    resp = await client.post(
        "/api/v1/exports", headers=headers, json={"format": "json", "columns": ["company"]}
    )
    payload = json.loads((await client.get(resp.json()["download_url"], headers=headers)).content)
    assert set(payload["rows"][0]) == {"company"}
    assert "email" not in payload["rows"][0]


# --- Resources ------------------------------------------------------------


async def test_search_results_export(client, signed_up_user, google_places_configured):
    _, headers = signed_up_user
    search = await client.post(
        "/api/v1/search", headers=headers, json={"query": "switchgear", "location": "Pune"}
    )
    assert search.status_code == 201
    search_id = search.json()["id"]

    resp = await client.post(
        "/api/v1/exports",
        headers=headers,
        json={"resource": "search_results", "format": "csv", "search_id": search_id},
    )
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["resource"] == "search_results"
    assert body["row_count"] == 2

    blob = (await client.get(body["download_url"], headers=headers)).content
    text = blob.decode("utf-8-sig")
    assert "switchgear" in text  # the query is recorded in the file's metadata
    assert "Apex Switchgear Pvt Ltd" in text


async def test_search_results_requires_a_search_id(client, signed_up_user):
    _, headers = signed_up_user
    resp = await client.post(
        "/api/v1/exports", headers=headers, json={"resource": "search_results", "format": "csv"}
    )
    assert resp.status_code == 400
    assert "search_id" in resp.json()["message"]


async def test_search_results_rejects_an_unknown_search(client, signed_up_user):
    _, headers = signed_up_user
    resp = await client.post(
        "/api/v1/exports",
        headers=headers,
        json={"resource": "search_results", "format": "csv", "search_id": str(uuid.uuid4())},
    )
    assert resp.status_code == 404


@pytest.mark.parametrize("resource", ["dashboard_report", "analytics_report"])
async def test_report_exports_build_multi_section_documents(client, seeded_leads, resource):
    headers, _ = seeded_leads

    resp = await client.post("/api/v1/exports", headers=headers, json={"resource": resource, "format": "excel"})
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["status"] == "ready"
    assert body["resource"] == resource
    assert body["row_count"] > 0

    blob = (await client.get(body["download_url"], headers=headers)).content
    workbook = load_workbook(io.BytesIO(blob))
    # One sheet per report section — the reason XLSX is the best format for these.
    assert len(workbook.sheetnames) >= 3


@pytest.mark.parametrize("resource", ["dashboard_report", "analytics_report"])
async def test_reports_render_as_pdf(client, seeded_leads, resource):
    headers, _ = seeded_leads
    resp = await client.post("/api/v1/exports", headers=headers, json={"resource": resource, "format": "pdf"})
    assert resp.status_code == 201, resp.text
    blob = (await client.get(resp.json()["download_url"], headers=headers)).content
    assert blob.startswith(b"%PDF-")


async def test_reports_are_always_generated_inline(client, seeded_leads, monkeypatch):
    """Reports are async-only aggregates, so they must never be queued."""
    from config.settings import settings

    monkeypatch.setattr(settings, "EXPORT_ASYNC_ROW_THRESHOLD", 1, raising=False)
    headers, _ = seeded_leads

    resp = await client.post(
        "/api/v1/exports", headers=headers, json={"resource": "dashboard_report", "format": "csv"}
    )
    assert resp.status_code == 201
    assert resp.json()["status"] == "ready"


# --- Security: tenant isolation ------------------------------------------


@pytest.fixture
async def two_orgs(client: AsyncClient, signed_up_user):
    """(owner_headers_with_an_export, other_org_headers)"""
    _, headers = signed_up_user
    await _create_leads(client, headers, [LEAD_SPECS[0]])
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    assert created.status_code == 201

    other = await client.post(
        "/api/v1/auth/signup",
        json={
            "email": f"outsider_{uuid.uuid4().hex[:8]}@example.com", "password": "TestPass123",
            "full_name": "Outsider", "company_name": "Outsider Co",
        },
    )
    return headers, {"Authorization": f"Bearer {other.json()['access_token']}"}, created.json()


async def test_another_org_cannot_read_the_export(client, two_orgs):
    _, outsider, export = two_orgs
    resp = await client.get(f"/api/v1/exports/{export['id']}", headers=outsider)
    # 404 rather than 403: a 403 would confirm the id exists.
    assert resp.status_code == 404


async def test_another_org_cannot_download_the_export(client, two_orgs):
    _, outsider, export = two_orgs
    resp = await client.get(export["download_url"], headers=outsider)
    assert resp.status_code == 404


async def test_another_org_cannot_delete_the_export(client, two_orgs):
    _, outsider, export = two_orgs
    resp = await client.delete(f"/api/v1/exports/{export['id']}", headers=outsider)
    assert resp.status_code == 404


async def test_another_org_cannot_mint_a_download_token(client, two_orgs):
    _, outsider, export = two_orgs
    resp = await client.post(f"/api/v1/exports/{export['id']}/download-token", headers=outsider)
    assert resp.status_code == 404


async def test_history_is_scoped_to_the_organization(client, two_orgs):
    _, outsider, _export = two_orgs
    resp = await client.get("/api/v1/exports", headers=outsider)
    assert resp.json()["meta"]["total_items"] == 0


async def test_export_requires_authentication(client):
    assert (await client.post("/api/v1/exports", json={"format": "csv"})).status_code in (401, 403)
    assert (await client.get("/api/v1/exports")).status_code in (401, 403)


# --- Security: RBAC -------------------------------------------------------


@pytest.fixture
async def viewer_headers(client: AsyncClient, signed_up_user, db_session):
    """A second user joined to the same organization with the VIEWER role.

    VIEWER is the one seeded role without `leads.export`, which is what makes it
    the meaningful negative case.
    """
    _, owner_headers = signed_up_user
    org_id = (
        await db_session.execute(select(Organization.id).order_by(Organization.created_at.desc()).limit(1))
    ).scalar_one()

    email = f"viewer_{uuid.uuid4().hex[:8]}@example.com"
    signup = await client.post(
        "/api/v1/auth/signup",
        json={"email": email, "password": "TestPass123", "full_name": "Viewer", "company_name": "Viewer Own Co"},
    )
    token = signup.json()["access_token"]

    from models.user import User

    user_id = (await db_session.execute(select(User.id).where(User.email == email))).scalar_one()
    viewer_role_id = (
        await db_session.execute(select(Role.id).where(Role.name == RoleName.VIEWER))
    ).scalar_one()

    db_session.add(
        OrganizationMember(
            organization_id=org_id, user_id=user_id, role_id=viewer_role_id, status="active"
        )
    )
    await db_session.commit()

    # X-Organization-Id targets the shared org rather than the viewer's own.
    return {"Authorization": f"Bearer {token}", "X-Organization-Id": str(org_id)}, owner_headers


async def test_viewer_cannot_create_an_export(client, viewer_headers):
    viewer, _owner = viewer_headers
    resp = await client.post("/api/v1/exports", headers=viewer, json={"format": "csv"})
    assert resp.status_code == 403
    assert "leads.export" in resp.json()["message"]


async def test_viewer_cannot_delete_an_export(client, viewer_headers, seeded_leads):
    viewer, owner = viewer_headers
    created = await client.post("/api/v1/exports", headers=owner, json={"format": "csv"})
    resp = await client.delete(f"/api/v1/exports/{created.json()['id']}", headers=viewer)
    assert resp.status_code == 403


async def test_viewer_can_still_read_history(client, viewer_headers):
    """Seeing that an export happened is not the same as extracting data."""
    viewer, _owner = viewer_headers
    assert (await client.get("/api/v1/exports", headers=viewer)).status_code == 200


async def test_member_role_can_export(client, signed_up_user, db_session):
    """MEMBER holds leads.export in the seeded mapping, so it must be allowed."""
    _, owner_headers = signed_up_user
    org_id = (
        await db_session.execute(select(Organization.id).order_by(Organization.created_at.desc()).limit(1))
    ).scalar_one()

    email = f"member_{uuid.uuid4().hex[:8]}@example.com"
    signup = await client.post(
        "/api/v1/auth/signup",
        json={"email": email, "password": "TestPass123", "full_name": "Member", "company_name": "Member Own Co"},
    )
    from models.user import User

    user_id = (await db_session.execute(select(User.id).where(User.email == email))).scalar_one()
    member_role_id = (await db_session.execute(select(Role.id).where(Role.name == RoleName.MEMBER))).scalar_one()
    db_session.add(
        OrganizationMember(organization_id=org_id, user_id=user_id, role_id=member_role_id, status="active")
    )
    await db_session.commit()

    headers = {"Authorization": f"Bearer {signup.json()['access_token']}", "X-Organization-Id": str(org_id)}
    assert (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).status_code == 201


# --- Security: download tokens -------------------------------------------


async def test_signed_token_downloads_without_an_auth_header(client, seeded_leads):
    """A browser <a href> cannot set headers; this is the path that makes it work."""
    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    export_id = created.json()["id"]

    minted = await client.post(f"/api/v1/exports/{export_id}/download-token", headers=headers)
    assert minted.status_code == 200
    body = minted.json()
    assert body["expires_in"] > 0

    # No Authorization header at all.
    download = await client.get(body["download_url"])
    assert download.status_code == 200
    assert download.content.startswith(b"\xef\xbb\xbf")


async def test_download_without_any_credential_is_rejected(client, seeded_leads):
    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    resp = await client.get(created.json()["download_url"])
    assert resp.status_code == 401


async def test_a_token_for_one_export_cannot_download_another(client, seeded_leads):
    headers, _ = seeded_leads
    first = (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).json()
    second = (await client.post("/api/v1/exports", headers=headers, json={"format": "json"})).json()

    token = (await client.post(f"/api/v1/exports/{first['id']}/download-token", headers=headers)).json()["token"]

    resp = await client.get(f"/api/v1/exports/{second['id']}/download", params={"token": token})
    assert resp.status_code == 401


@pytest.mark.parametrize("bad_token", ["", "garbage", "a.b", "x" * 600])
async def test_malformed_tokens_are_rejected(client, seeded_leads, bad_token):
    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    resp = await client.get(created.json()["download_url"], params={"token": bad_token})
    assert resp.status_code == 401


async def test_a_tampered_token_is_rejected(client, seeded_leads):
    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    export_id = created.json()["id"]
    token = (await client.post(f"/api/v1/exports/{export_id}/download-token", headers=headers)).json()["token"]

    flipped = token[:-1] + ("A" if token[-1] != "A" else "B")
    resp = await client.get(f"/api/v1/exports/{export_id}/download", params={"token": flipped})
    assert resp.status_code == 401


async def test_an_expired_token_is_rejected(client, seeded_leads):
    from utils.download_token import issue

    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    export_id = created.json()["id"]

    token, _ttl = issue(uuid.UUID(export_id), uuid.uuid4(), ttl_seconds=1)
    time.sleep(1.2)
    resp = await client.get(f"/api/v1/exports/{export_id}/download", params={"token": token})
    assert resp.status_code == 401


async def test_a_token_from_a_removed_member_stops_working(client, seeded_leads, db_session):
    """Membership is re-checked at download time, not trusted from the token."""
    from utils.download_token import issue

    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    export_id = created.json()["id"]

    # A user who was never a member stands in for one removed after minting.
    token, _ = issue(uuid.UUID(export_id), uuid.uuid4())
    resp = await client.get(f"/api/v1/exports/{export_id}/download", params={"token": token})
    assert resp.status_code == 404


async def test_download_count_is_audited(client, seeded_leads):
    headers, _ = seeded_leads
    created = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    assert created.json()["download_count"] == 0

    for _ in range(3):
        await client.get(created.json()["download_url"], headers=headers)

    fetched = await client.get(f"/api/v1/exports/{created.json()['id']}", headers=headers)
    assert fetched.json()["download_count"] == 3


# --- Security: rate limiting ---------------------------------------------


async def test_export_creation_is_rate_limited_per_user(client, seeded_leads, monkeypatch):
    from config.settings import settings

    monkeypatch.setattr(settings, "EXPORT_RATE_LIMIT_PER_HOUR", 3, raising=False)
    headers, _ = seeded_leads

    statuses = []
    for _ in range(5):
        resp = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
        statuses.append(resp.status_code)

    assert statuses[:3] == [201, 201, 201]
    assert 429 in statuses[3:]


async def test_rate_limit_does_not_block_reads(client, seeded_leads, monkeypatch):
    """Only creation is budgeted; history and downloads must stay available."""
    from config.settings import settings

    monkeypatch.setattr(settings, "EXPORT_RATE_LIMIT_PER_HOUR", 1, raising=False)
    headers, _ = seeded_leads

    created = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    assert created.status_code == 201
    assert (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).status_code == 429

    assert (await client.get("/api/v1/exports", headers=headers)).status_code == 200
    assert (await client.get(created.json()["download_url"], headers=headers)).status_code == 200


# --- Security: size and row limits ---------------------------------------


async def test_row_cap_refuses_before_generating(client, seeded_leads, monkeypatch, db_session):
    from config.settings import settings

    monkeypatch.setattr(settings, "EXPORT_MAX_ROWS", 2, raising=False)
    headers, _ = seeded_leads

    resp = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    assert resp.status_code == 400
    assert "3" in resp.json()["message"] and "2" in resp.json()["message"]

    # Refused before rendering, so no file and no export row were created.
    assert (await db_session.execute(select(Export))).scalars().all() == []


async def test_size_cap_rejects_and_records_the_failure(client, seeded_leads, monkeypatch, db_session):
    from config.settings import settings

    monkeypatch.setattr(settings, "EXPORT_MAX_FILE_SIZE_MB", 0, raising=False)
    headers, _ = seeded_leads

    resp = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    assert resp.status_code == 400
    assert "limit" in resp.json()["message"].lower()

    # The attempt is part of the audit trail, with a reason.
    rows = (await db_session.execute(select(Export))).scalars().all()
    assert len(rows) == 1
    assert rows[0].status is ExportStatus.FAILED
    assert rows[0].error_message


async def test_a_failed_export_has_no_download_url(client, seeded_leads, monkeypatch):
    from config.settings import settings

    monkeypatch.setattr(settings, "EXPORT_MAX_FILE_SIZE_MB", 0, raising=False)
    headers, _ = seeded_leads
    await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})

    history = await client.get("/api/v1/exports", headers=headers)
    failed = history.json()["items"][0]
    assert failed["status"] == "failed"
    assert failed["download_url"] is None


async def test_too_many_lead_ids_is_rejected_by_validation(client, seeded_leads):
    headers, _ = seeded_leads
    resp = await client.post(
        "/api/v1/exports",
        headers=headers,
        json={"format": "csv", "scope": "selected", "lead_ids": [str(uuid.uuid4()) for _ in range(10_001)]},
    )
    assert resp.status_code == 422


# --- Background processing ------------------------------------------------


async def test_large_exports_are_queued(client, seeded_leads, monkeypatch):
    from config.settings import settings

    monkeypatch.setattr(settings, "EXPORT_ASYNC_ROW_THRESHOLD", 2, raising=False)
    headers, _ = seeded_leads

    resp = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    assert resp.status_code == 202
    body = resp.json()
    assert body["status"] == "processing"
    assert body["download_url"] is None
    # The preflight count is reported so a client can show progress to a target.
    assert body["row_count"] == 3


async def test_downloading_a_queued_export_explains_itself(client, seeded_leads, monkeypatch):
    from config.settings import settings

    monkeypatch.setattr(settings, "EXPORT_ASYNC_ROW_THRESHOLD", 1, raising=False)
    headers, _ = seeded_leads

    queued = await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    resp = await client.get(f"/api/v1/exports/{queued.json()['id']}/download", headers=headers)
    assert resp.status_code == 400
    assert "still being generated" in resp.json()["message"]


async def test_the_worker_task_generates_a_queued_export(seeded_leads, monkeypatch, db_session, client):
    """Runs the Celery task body directly against the sync engine.

    Calling the task function rather than dispatching it exercises the real sync
    query path — the half of the codebase a broker-less test would never touch.
    """
    from config.settings import settings

    monkeypatch.setattr(settings, "EXPORT_ASYNC_ROW_THRESHOLD", 1, raising=False)
    headers, _ = seeded_leads

    queued = (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).json()
    assert queued["status"] == "processing"

    from services.export_tasks import generate_export_task

    result = generate_export_task(queued["id"])
    assert result["status"] == "ready", result
    assert result["rows"] == 3

    fetched = await client.get(f"/api/v1/exports/{queued['id']}", headers=headers)
    body = fetched.json()
    assert body["status"] == "ready"
    assert body["size_bytes"] > 0
    assert body["download_url"] is not None

    download = await client.get(body["download_url"], headers=headers)
    assert download.status_code == 200
    rows = list(csv.reader(io.StringIO(download.content.decode("utf-8-sig"))))
    assert any(r and r[0] == "Apex Switchgear Pvt Ltd" for r in rows)


async def test_the_worker_task_is_idempotent(seeded_leads, monkeypatch, client):
    """Celery delivers at least once; a redelivery must not produce a second file."""
    from config.settings import settings

    monkeypatch.setattr(settings, "EXPORT_ASYNC_ROW_THRESHOLD", 1, raising=False)
    headers, _ = seeded_leads
    queued = (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).json()

    from services.export_tasks import generate_export_task

    assert generate_export_task(queued["id"])["status"] == "ready"
    assert generate_export_task(queued["id"])["status"] == "ready"  # second delivery: no-op


async def test_the_worker_task_tolerates_a_missing_row(monkeypatch, tmp_path):
    from services.export_tasks import generate_export_task

    assert generate_export_task(str(uuid.uuid4()))["status"] == "missing"


# --- Expiry and cleanup ---------------------------------------------------


async def test_an_expired_export_cannot_be_downloaded(client, seeded_leads, db_session):
    headers, _ = seeded_leads
    created = (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).json()

    export = (await db_session.execute(select(Export).where(Export.id == uuid.UUID(created["id"])))).scalar_one()
    export.expires_at = datetime.now(UTC) - timedelta(hours=1)
    await db_session.commit()

    resp = await client.get(created["download_url"], headers=headers)
    assert resp.status_code == 404
    assert "expired" in resp.json()["message"].lower()


async def test_an_expired_export_has_no_download_url(client, seeded_leads, db_session):
    headers, _ = seeded_leads
    created = (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).json()

    export = (await db_session.execute(select(Export).where(Export.id == uuid.UUID(created["id"])))).scalar_one()
    export.expires_at = datetime.now(UTC) - timedelta(hours=1)
    await db_session.commit()

    fetched = await client.get(f"/api/v1/exports/{created['id']}", headers=headers)
    assert fetched.json()["download_url"] is None


async def test_purge_deletes_files_and_keeps_the_audit_row(client, seeded_leads, db_session):
    """Bytes go, history stays — who extracted what is worth more than the file."""
    from pathlib import Path

    from services import export_service

    headers, _ = seeded_leads
    created = (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).json()

    export = (await db_session.execute(select(Export).where(Export.id == uuid.UUID(created["id"])))).scalar_one()
    stored = Path(export.storage_path)
    assert stored.exists()

    export.expires_at = datetime.now(UTC) - timedelta(hours=1)
    await db_session.commit()

    result = await export_service.purge_expired_exports(db_session)
    assert result == {"expired": 1, "files_deleted": 1}
    assert not stored.exists()

    await db_session.refresh(export)
    assert export.status is ExportStatus.EXPIRED
    assert export.storage_path is None

    # The row is still listed, so history remains a complete record.
    history = await client.get("/api/v1/exports", headers=headers)
    assert history.json()["meta"]["total_items"] == 1
    assert history.json()["items"][0]["status"] == "expired"


async def test_purge_is_idempotent(client, seeded_leads, db_session):
    from services import export_service

    headers, _ = seeded_leads
    created = (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).json()
    export = (await db_session.execute(select(Export).where(Export.id == uuid.UUID(created["id"])))).scalar_one()
    export.expires_at = datetime.now(UTC) - timedelta(hours=1)
    await db_session.commit()

    assert (await export_service.purge_expired_exports(db_session))["expired"] == 1
    assert (await export_service.purge_expired_exports(db_session))["expired"] == 0


async def test_purge_leaves_live_exports_alone(client, seeded_leads, db_session):
    from services import export_service

    headers, _ = seeded_leads
    created = (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).json()

    assert (await export_service.purge_expired_exports(db_session))["expired"] == 0
    assert (await client.get(created["download_url"], headers=headers)).status_code == 200


# --- Delete ---------------------------------------------------------------


async def test_delete_removes_the_row_and_the_file(client, seeded_leads, db_session):
    from pathlib import Path

    headers, _ = seeded_leads
    created = (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).json()
    export = (await db_session.execute(select(Export).where(Export.id == uuid.UUID(created["id"])))).scalar_one()
    stored = Path(export.storage_path)
    assert stored.exists()

    resp = await client.delete(f"/api/v1/exports/{created['id']}", headers=headers)
    assert resp.status_code == 200

    assert not stored.exists()
    assert (await client.get(f"/api/v1/exports/{created['id']}", headers=headers)).status_code == 404


async def test_deleting_an_unknown_export_is_404(client, signed_up_user):
    _, headers = signed_up_user
    assert (await client.delete(f"/api/v1/exports/{uuid.uuid4()}", headers=headers)).status_code == 404


# --- History --------------------------------------------------------------


async def test_history_is_newest_first_and_paginated(client, seeded_leads):
    headers, _ = seeded_leads
    for fmt in ("csv", "json", "excel", "pdf"):
        await client.post("/api/v1/exports", headers=headers, json={"format": fmt})

    page = await client.get("/api/v1/exports", headers=headers, params={"page": 1, "page_size": 2})
    body = page.json()
    assert body["meta"]["total_items"] == 4
    assert body["meta"]["total_pages"] == 2
    assert body["meta"]["has_next"] is True
    assert len(body["items"]) == 2
    # Newest first: pdf was created last.
    assert body["items"][0]["format"] == "pdf"

    second = await client.get("/api/v1/exports", headers=headers, params={"page": 2, "page_size": 2})
    assert second.json()["meta"]["has_previous"] is True
    assert len({i["id"] for i in body["items"]} & {i["id"] for i in second.json()["items"]}) == 0


async def test_history_can_be_filtered(client, seeded_leads):
    headers, _ = seeded_leads
    await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})
    await client.post("/api/v1/exports", headers=headers, json={"resource": "dashboard_report", "format": "csv"})

    by_resource = await client.get("/api/v1/exports", headers=headers, params={"resource": "dashboard_report"})
    assert by_resource.json()["meta"]["total_items"] == 1

    by_status = await client.get("/api/v1/exports", headers=headers, params={"status": "ready"})
    assert by_status.json()["meta"]["total_items"] == 2


# --- Options endpoint -----------------------------------------------------


async def test_options_endpoint_describes_the_module(client, signed_up_user):
    _, headers = signed_up_user
    resp = await client.get("/api/v1/exports/formats", headers=headers)
    assert resp.status_code == 200
    body = resp.json()

    assert {f["value"] for f in body["formats"]} == {"csv", "excel", "pdf", "json"}
    assert next(f for f in body["formats"] if f["value"] == "excel")["extension"] == "xlsx"
    assert set(body["resources"]) == {"leads", "search_results", "dashboard_report", "analytics_report"}
    assert "lead_score" in {c["key"] for c in body["lead_columns"]}
    assert body["limits"]["max_rows"] > 0


# --- Data safety ----------------------------------------------------------


async def test_spreadsheet_formula_injection_is_neutralized(client, signed_up_user):
    """Lead data is attacker-influenced, so an exported cell must not execute.

    A company name arriving from a third-party provider or a user's CSV import can
    start with '='; written raw, the recipient's spreadsheet evaluates it on open.
    """
    _, headers = signed_up_user
    await _create_leads(client, headers, [{"company": "=cmd|'/c calc'!A1", "city": "Pune"}])

    created = (await client.post("/api/v1/exports", headers=headers, json={"format": "csv"})).json()
    blob = (await client.get(created["download_url"], headers=headers)).content
    rows = list(csv.reader(io.StringIO(blob.decode("utf-8-sig"))))
    company_cell = next(r[0] for r in rows if r and r[0].endswith("!A1"))
    assert company_cell.startswith("'="), company_cell

    xlsx = (await client.post("/api/v1/exports", headers=headers, json={"format": "excel"})).json()
    wb = load_workbook(io.BytesIO((await client.get(xlsx["download_url"], headers=headers)).content))
    cells = [c.value for row in wb.active.iter_rows() for c in row if isinstance(c.value, str)]
    assert any(v.startswith("'=") for v in cells)


async def test_names_with_markup_do_not_break_the_pdf(client, signed_up_user):
    """ReportLab parses inline markup, so unescaped <, > or & fails the build."""
    _, headers = signed_up_user
    await _create_leads(client, headers, [{"company": "Smith & Sons <Holdings> Ltd", "city": "Pune"}])

    resp = await client.post("/api/v1/exports", headers=headers, json={"format": "pdf"})
    assert resp.status_code == 201, resp.text
    blob = (await client.get(resp.json()["download_url"], headers=headers)).content
    assert blob.startswith(b"%PDF-")


async def test_file_name_extension_always_matches_the_bytes(client, seeded_leads):
    """A file named .csv must never actually be a workbook."""
    headers, _ = seeded_leads
    resp = await client.post(
        "/api/v1/exports", headers=headers, json={"format": "excel", "file_name": "report.csv"}
    )
    assert resp.json()["file_name"] == "report.xlsx"


async def test_path_traversal_in_the_file_name_is_neutralized(client, seeded_leads, db_session):
    headers, _ = seeded_leads
    resp = await client.post(
        "/api/v1/exports", headers=headers, json={"format": "csv", "file_name": "../../../etc/passwd"}
    )
    assert resp.status_code == 201

    export = (
        await db_session.execute(select(Export).where(Export.id == uuid.UUID(resp.json()["id"])))
    ).scalar_one()
    assert ".." not in export.storage_path
    # The file landed under the organization's own export prefix.
    assert f"{export.organization_id}" in export.storage_path
    assert "exports" in export.storage_path


# Size-label formatting and the other pure, database-free behaviour of the
# writers live in tests/test_export_formats.py — this module is the HTTP-level
# integration suite, and its session-scoped asyncio mark applies to every test in
# it, which a synchronous test cannot carry.


# --- Empty data -----------------------------------------------------------


@pytest.mark.parametrize("fmt", ["csv", "excel", "pdf", "json"])
async def test_exporting_with_no_leads_still_produces_a_valid_file(client, signed_up_user, fmt):
    """An empty result is a valid export, not an error."""
    _, headers = signed_up_user
    resp = await client.post("/api/v1/exports", headers=headers, json={"format": fmt})
    assert resp.status_code == 201, resp.text
    assert resp.json()["row_count"] == 0

    download = await client.get(resp.json()["download_url"], headers=headers)
    assert download.status_code == 200
    assert len(download.content) > 0
